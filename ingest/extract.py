import io
import os
import subprocess
import tempfile
import logging
import fitz  # PyMuPDF

logger = logging.getLogger(__name__)


def classify_page(page):
    text = page.get_text().strip()
    text_len = len(text)
    images = page.get_images()
    drawings = page.get_drawings()
    page_area = abs(page.rect)

    if text_len > 100 and not images:
        return "text"
    if text_len > 100 and images:
        return "mixed"
    if not text_len and images:
        for blk in page.get_text("dict")["blocks"]:
            if blk.get("type") == 1:
                bbox = fitz.Rect(blk["bbox"])
                if page_area > 0 and abs(bbox & page.rect) / page_area >= 0.95:
                    return "scanned"
        return "image-heavy"
    if len(drawings) > 200 and not images:
        return "vector-drawing"
    if text_len > 0 and text_len < 100:
        return "sparse"
    if text_len > 0:
        return "text"
    return "unknown"


def ocr_page_bytes(page_pdf_bytes, dpi=300):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as inp, \
         tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as out:
        inp.write(page_pdf_bytes)
        inp.flush()
        inp_path = inp.name
        out_path = out.name

    try:
        result = subprocess.run(
            [
                "ocrmypdf",
                "--jobs", "2",
                "--output-type", "pdf",
                "--skip-text",
                "--max-image-mpixels", "500",
                "--skip-big", "80",
                "--image-dpi", str(dpi),
                "--quiet",
                inp_path,
                out_path,
            ],
            capture_output=True,
            text=True,
            timeout=300,
        )
        if result.returncode not in (0, 6):  # 6 = already has text
            logger.warning(f"OCRmyPDF returned {result.returncode}: {result.stderr[:200]}")
            return ""

        doc = fitz.open(out_path)
        text = "\n".join(p.get_text().strip() for p in doc if p.get_text().strip())
        doc.close()
        return text
    except subprocess.TimeoutExpired:
        logger.warning("OCRmyPDF timeout")
        return ""
    except Exception as e:
        logger.warning(f"OCR failed: {e}")
        return ""
    finally:
        for p in (inp_path, out_path):
            try:
                os.unlink(p)
            except OSError:
                pass


def extract_single_page_pdf(doc, page_num):
    new_doc = fitz.open()
    new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
    pdf_bytes = new_doc.tobytes()
    new_doc.close()
    return pdf_bytes


def extract_pdf_text(file_bytes):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    all_text = []
    total_pages = len(doc)
    ocr_count = 0

    for i, page in enumerate(doc):
        page_type = classify_page(page)

        if page_type == "text":
            text = page.get_text().strip()
            all_text.append(text)

        elif page_type == "mixed":
            text = page.get_text().strip()
            page_pdf = extract_single_page_pdf(doc, i)
            ocr_text = ocr_page_bytes(page_pdf, dpi=300)
            if ocr_text and len(ocr_text.split()) > len(text.split()) + 10:
                all_text.append(ocr_text)
                ocr_count += 1
            else:
                all_text.append(text)

        elif page_type == "scanned":
            page_pdf = extract_single_page_pdf(doc, i)
            ocr_text = ocr_page_bytes(page_pdf, dpi=300)
            if ocr_text:
                all_text.append(ocr_text)
                ocr_count += 1
            else:
                logger.info(f"Page {i+1}: scanned but OCR returned nothing")

        elif page_type == "image-heavy":
            page_pdf = extract_single_page_pdf(doc, i)
            ocr_text = ocr_page_bytes(page_pdf, dpi=200)
            if ocr_text:
                all_text.append(ocr_text)
                ocr_count += 1

        elif page_type == "sparse":
            page_pdf = extract_single_page_pdf(doc, i)
            ocr_text = ocr_page_bytes(page_pdf, dpi=300)
            if ocr_text and len(ocr_text.split()) > len(page.get_text().strip().split()):
                all_text.append(ocr_text)
                ocr_count += 1
                logger.info(f"Page {i+1}: sparse text, OCR got {len(ocr_text.split())} words")
            else:
                text = page.get_text().strip()
                if text:
                    all_text.append(text)

        elif page_type == "vector-drawing":
            text = page.get_text().strip()
            if text:
                all_text.append(text)

        else:
            logger.info(f"Page {i+1}: unknown type, skipping")

    doc.close()
    result = "\n\n".join(all_text)
    logger.info(f"PDF: {total_pages} pages, {ocr_count} OCR'd, {len(result.split())} words")
    return result.strip()


def extract_email_text(raw_email):
    body = raw_email.get('body', '')
    if raw_email.get('html'):
        import html2text
        h = html2text.HTML2Text()
        h.ignore_links = True
        h.ignore_images = True
        body = h.handle(raw_email['html'])

    # Extract text from attachments
    for att in raw_email.get('attachments', []):
        if not att.get('data'):
            continue
        mime = att.get('mime_type', '').lower()
        filename = att.get('filename', '').lower()
        att_text = ''

        try:
            if 'pdf' in mime or filename.endswith('.pdf'):
                att_text = extract_pdf_text(att['data'])
            elif 'word' in mime or filename.endswith('.docx') or filename.endswith('.doc'):
                att_text = extract_docx_text(att['data'])
            elif 'spreadsheet' in mime or filename.endswith('.xlsx') or filename.endswith('.xls'):
                att_text = extract_xlsx_text(att['data'])
            elif 'text' in mime or filename.endswith('.txt') or filename.endswith('.csv'):
                att_text = att['data'].decode('utf-8', errors='ignore')
            else:
                logger.info(f"Skipping unsupported attachment: {att.get('filename', 'unknown')} ({mime})")
                continue

            if att_text:
                logger.info(f"Attachment extracted: {att.get('filename', 'unknown')} ({len(att_text.split())} words)")
                body += f"\n\n--- Attachment: {att.get('filename', 'unknown')} ---\n{att_text}"
        except Exception as e:
            logger.warning(f"Failed to extract attachment {att.get('filename', 'unknown')}: {e}")

    return body.strip()


def extract_docx_text(file_bytes):
    import docx
    doc = docx.Document(io.BytesIO(file_bytes))
    return '\n'.join(para.text for para in doc.paragraphs if para.text)


def extract_xlsx_text(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows.append(f"--- Sheet: {sheet} ---")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else '' for c in row]
            if any(vals):
                rows.append('\t'.join(vals))
    wb.close()
    return '\n'.join(rows)


def extract_text(raw_input, file_bytes=None):
    if 'body' in raw_input or 'html' in raw_input:
        return extract_email_text(raw_input)

    file_type = raw_input.get('file_type', '').lower()

    if 'pdf' in file_type:
        return extract_pdf_text(file_bytes)
    elif 'docx' in file_type or 'word' in file_type:
        return extract_docx_text(file_bytes)
    elif 'spreadsheet' in file_type or 'xlsx' in file_type or 'excel' in file_type:
        return extract_xlsx_text(file_bytes)
    elif file_bytes:
        return file_bytes.decode('utf-8', errors='ignore')
    return ''
